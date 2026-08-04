"""Builds the daily 5-sheet Excel report."""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config import settings
from src.utils.logger import log

_ACTION_FILL = {
    "BUY NOW": "C6EFCE",
    "BUY": "DDEBF7",
    "WATCH": "FFF2CC",
    "AVOID": "FCE4D6",
    "SKIP": "F2F2F2",
}
_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(bold=True, color="FFFFFF")


def _session_label() -> str:
    """Classify the current run by IST market-session time."""
    from datetime import timezone, timedelta
    ist = datetime.now(timezone(timedelta(hours=5, minutes=30)))
    if ist.hour < 9 or (ist.hour == 9 and ist.minute < 15):
        return "Pre-Market"
    if ist.hour < 15 or (ist.hour == 15 and ist.minute < 30):
        return "Mid-Session"
    return "Post-Close"


def _trade_levels(row: pd.Series) -> dict:
    entry = row["close"]
    stop = round(entry * (1 - settings.STOP_PCT), 2)
    risk = entry - stop
    return {
        "entry": entry,
        "stop": stop,
        "target1": round(entry + settings.R_R_T1 * risk, 2),
        "target2": round(entry + settings.R_R_T2 * risk, 2),
        "rr": settings.R_R_T2,
    }


def _autosize(ws):
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(width + 2, 40)


def _style_header(ws):
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _write_df(writer, df: pd.DataFrame, sheet: str):
    df.to_excel(writer, sheet_name=sheet, index=False)
    ws = writer.sheets[sheet]
    _style_header(ws)
    _autosize(ws)
    return ws


def _color_actions(ws, action_col_idx: int):
    for r in range(2, ws.max_row + 1):
        action = ws.cell(row=r, column=action_col_idx).value
        fill = _ACTION_FILL.get(action)
        if fill:
            ws.cell(row=r, column=action_col_idx).fill = PatternFill("solid", fgColor=fill)


def generate_report(df: pd.DataFrame, out_dir: Path | None = None) -> Path:
    out_dir = Path(out_dir or settings.OUTPUT_DIR)
    out_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y-%m-%d")
    session = _session_label()
    path = out_dir / f"breakout_scan_{stamp}_{session}.xlsx"

    master_cols = [
        "true_rank", "symbol", "close", "breakout_level", "pct_above", "rsi",
        "volume_x", "turnover_cr", "adx", "macd_hist", "rel_strength", "follow_through",
        "setup_type", "live_status", "final_score", "action", "position_size",
        "sector", "market_cap", "pe_ratio", "roe", "debt_to_equity", "earnings_growth",
        "fundamental_bonus",
    ]
    master = df[[c for c in master_cols if c in df.columns]].copy()

    actionable = df[df["action"].isin(["BUY NOW", "BUY"])].copy()
    trade_rows = []
    for _, row in actionable.iterrows():
        lv = _trade_levels(row)
        trade_rows.append({
            "true_rank": row["true_rank"], "symbol": row["symbol"],
            "action": row["action"], "final_score": row["final_score"],
            "entry": lv["entry"], "stop": lv["stop"],
            "target1": lv["target1"], "target2": lv["target2"],
            "R:R": lv["rr"], "position_size": row["position_size"],
        })
    trades = pd.DataFrame(trade_rows)

    comp_cols = ["true_rank", "symbol", "entry_checks", "trend_alignment", "momentum",
                 "rsi_health", "proximity", "liquidity", "live_score", "volume",
                 "penalty", "final_score"]
    breakdown = df[[c for c in comp_cols if c in df.columns]].copy()

    checklist = pd.DataFrame({
        "#": range(1, 11),
        "Pre-Market Check (before 9:15 AM)": [
            "Confirm index (NIFTY/SENSEX) not gapping down >0.5%",
            "Re-verify live price is still near breakout level",
            "Check for overnight news / results on the stock",
            "Confirm volume interest in pre-open session",
            "Verify no upper-circuit / illiquidity risk",
            "Set stop-loss order at planned level",
            "Size position per capital allocation band",
            "Note target1 / target2 exits",
            "Check sector is not broadly weak today",
            "Log the trade plan before entering",
        ],
        "Done": [""] * 10,
    })

    fundamental_cols_present = [c for c in [
        "sector", "market_cap", "pe_ratio", "roe", "debt_to_equity", "earnings_growth",
    ] if c in df.columns]
    if fundamental_cols_present:
        has_fundamentals = df[fundamental_cols_present].notna().any(axis=1)
    else:
        has_fundamentals = pd.Series([False] * len(df), index=df.index)
    top_picks_cols = (["true_rank", "symbol", "action", "final_score", "close",
                       "follow_through", "adx", "macd_hist"] + fundamental_cols_present)
    top_picks = df[has_fundamentals][[c for c in top_picks_cols if c in df.columns]].copy()
    top_picks = top_picks.sort_values("final_score", ascending=False).head(settings.TOP_PICKS_N)

    counts = df["action"].value_counts().to_dict()
    summary = pd.DataFrame({
        "Metric": [
            "Scan date", "Total screened",
            "BUY NOW", "BUY", "WATCH", "AVOID", "SKIP",
            "Top symbol", "Top score", "Scan session", "Top picks",
        ],
        "Value": [
            stamp, len(df),
            counts.get("BUY NOW", 0), counts.get("BUY", 0),
            counts.get("WATCH", 0), counts.get("AVOID", 0),
            counts.get("SKIP", 0),
            df.iloc[0]["symbol"] if len(df) else "-",
            df.iloc[0]["final_score"] if len(df) else "-",
            session, len(top_picks),
        ],
    })

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        ws_master = _write_df(writer, master, "Master Scan")
        if "action" in master.columns:
            _color_actions(ws_master, list(master.columns).index("action") + 1)

        _write_df(writer, trades if not trades.empty else pd.DataFrame(
            columns=["symbol", "action", "entry", "stop", "target1", "target2"]),
            "Actionable Trades")
        _write_df(writer, top_picks if not top_picks.empty else pd.DataFrame(columns=top_picks_cols), "Top Picks")
        _write_df(writer, breakdown, "Score Breakdown")
        _write_df(writer, checklist, "Pre-Market Checklist")
        _write_df(writer, summary, "Summary")

    log.info(f"Excel report written → {path}")
    return path
